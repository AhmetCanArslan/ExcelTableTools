import pandas as pd
import numpy as np
from typing import List, Dict, Any, Tuple, Optional
import threading
from queue import Queue
import time
import os
import gc
import psutil
from functools import lru_cache
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import math
from .preview_utils import apply_operation_to_partition

def calculate_optimal_chunk_size(file_size: int) -> int:
    """Calculate optimal chunk size based on file size and available memory."""
    available_memory = psutil.virtual_memory().available
    base_chunk_size = min(1000000, max(1000, file_size // 100))
    return min(base_chunk_size, available_memory // 10)  # Use at most 10% of available memory

class ChunkIterator:
    """Memory-efficient iterator for processing file chunks."""
    def __init__(self, file_path: str, chunk_size: int):
        self.file_path = file_path
        self.chunk_size = chunk_size
        self.total_rows = 0
        self._count_rows()
        self._iterator = None
    
    def _count_rows(self):
        """Count total rows efficiently."""
        if self.file_path.lower().endswith('.csv'):
            with open(self.file_path, 'rb') as f:
                self.total_rows = sum(1 for _ in f) - 1
        else:
            with openpyxl.load_workbook(self.file_path, read_only=True) as wb:
                self.total_rows = wb.active.max_row - 1

    def __iter__(self):
        """Initialize and return the iterator."""
        if self.file_path.lower().endswith('.csv'):
            self._iterator = pd.read_csv(
                self.file_path,
                chunksize=self.chunk_size,
                low_memory=False,
                dtype_backend='numpy_nullable',
                engine='c'
            )
        else:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            sheet = wb.active
            self.headers = [cell.value for cell in sheet[1]]
            self.current_chunk = []
            self.row_generator = sheet.iter_rows(min_row=2, values_only=True)
            self._iterator = self
            
        return self._iterator

    def __next__(self):
        """Return the next chunk of data."""
        if self.file_path.lower().endswith('.csv'):
            # For CSV files, delegate to pandas iterator
            return next(self._iterator)
        else:
            # For Excel files, handle chunking manually
            try:
                while len(self.current_chunk) < self.chunk_size:
                    row = next(self.row_generator)
                    self.current_chunk.append(row)
            except StopIteration:
                if not self.current_chunk:
                    raise StopIteration
            
            # Create DataFrame from accumulated rows
            chunk_df = pd.DataFrame(self.current_chunk, columns=self.headers)
            self.current_chunk = []  # Clear for next iteration
            return chunk_df

class DelayedOperationManager:
    def __init__(self):
        self.operations = []
        self.preview_df = None
        self.full_file_path = None
        self._cancel_flag = False
        self._progress_queue = Queue()
        self.input_file_type = None

    def _get_file_type(self, file_path: str) -> str:
        """Determine file type from extension."""
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            return 'csv'
        elif ext in ['.xls', '.xlsx']:
            return 'excel'
        else:
            raise ValueError(f"Unsupported file type: {ext}")

    @lru_cache(maxsize=32)
    def _get_column_metadata(self, column_name: str) -> Dict:
        """Cache column metadata to avoid recomputing."""
        return {
            'dtype': self.preview_df[column_name].dtype,
            'unique_count': len(self.preview_df[column_name].unique()),
            'has_nulls': self.preview_df[column_name].isnull().any()
        }

    def _optimize_dtypes(self, df: pd.DataFrame) -> pd.DataFrame:
        """Optimize DataFrame memory usage by choosing appropriate dtypes."""
        dtype_map = {
            'float64': 'float32',
            'int64': 'int32'
        }
        
        return df.astype({col: dtype_map.get(str(dtype), dtype) 
                         for col, dtype in df.dtypes.items()})

    def load_preview(self, file_path: str, position: str) -> pd.DataFrame:
        """Load a preview of the file based on position."""
        self.full_file_path = file_path
        self.input_file_type = self._get_file_type(file_path)
        
        nrows = 1000  # Fixed preview size
        
        try:
            if self.input_file_type == 'csv':
                total_rows = sum(1 for _ in open(file_path)) - 1
                if position == "head":
                    df = pd.read_csv(file_path, nrows=nrows, low_memory=False)
                elif position == "tail":
                    if total_rows <= nrows:
                        df = pd.read_csv(file_path, low_memory=False)
                    else:
                        skiprows = range(1, total_rows - nrows + 1)
                        df = pd.read_csv(file_path, skiprows=skiprows, low_memory=False)
                else:  # middle
                    if total_rows <= nrows:
                        df = pd.read_csv(file_path, low_memory=False)
                    else:
                        skiprows = range(1, (total_rows - nrows) // 2)
                        df = pd.read_csv(file_path, skiprows=skiprows, nrows=nrows, low_memory=False)
            else:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = wb.active
                total_rows = sheet.max_row - 1  # Exclude header

                if position == "head":
                    data = list(sheet.iter_rows(min_row=1, max_row=nrows+1, values_only=True))
                elif position == "tail":
                    if total_rows <= nrows:
                        data = list(sheet.iter_rows(values_only=True))
                    else:
                        start_row = total_rows - nrows + 1
                        data = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))  # Headers
                        data.extend(sheet.iter_rows(min_row=start_row, values_only=True))
                else:  # middle
                    if total_rows <= nrows:
                        data = list(sheet.iter_rows(values_only=True))
                    else:
                        middle_start = (total_rows - nrows) // 2
                        data = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))  # Headers
                        data.extend(sheet.iter_rows(min_row=middle_start, max_row=middle_start+nrows, values_only=True))

                headers = data[0]
                df = pd.DataFrame(data[1:], columns=headers)
                wb.close()

            # Optimize memory usage
            df = self._optimize_dtypes(df)
            return df

        except Exception as e:
            raise Exception(f"Error loading file: {str(e)}")

    def add_operation(self, operation: Dict[str, Any]):
        """Add a new operation to the queue."""
        if operation['type'] == 'column_operation':
            # Store the complete operation with all parameters
            op_copy = operation.copy()
            # Debug: Print operation details to ensure delimiter is preserved
            if 'delimiter' in op_copy:
                print(f"DEBUG: Adding operation with delimiter='{op_copy['delimiter']}'")
            self.operations.append(op_copy)

    def clear_operations(self):
        """Clear all pending operations and reset state."""
        self.operations = []
        self._cancel_flag = False
        self.full_file_path = None
        
        # Clear any cached data
        if hasattr(self, '_get_column_metadata'):
            self._get_column_metadata.cache_clear()
        
        # Force garbage collection
        gc.collect()

    def cancel_processing(self):
        """Cancel the current processing operation."""
        self._cancel_flag = True

    def _process_chunk(self, chunk: pd.DataFrame) -> pd.DataFrame:
        """Process a single chunk with all operations."""
        try:
            print(f"DEBUG: Starting to process chunk with {len(chunk)} rows")
            for i, op in enumerate(self.operations):
                if self._cancel_flag:
                    return None
                
                # Debug: Print operation details before processing
                if 'delimiter' in op:
                    print(f"DEBUG: Processing operation {i} with delimiter='{op['delimiter']}'")
                else:
                    print(f"DEBUG: Processing operation {i} of type '{op.get('key', 'unknown')}'")
                
                print(f"DEBUG: Chunk shape before operation {i}: {chunk.shape}")
                print(f"DEBUG: Chunk columns: {list(chunk.columns)}")
                
                try:
                    chunk = apply_operation_to_partition(chunk, op['type'], op)
                    print(f"DEBUG: Successfully completed operation {i}")
                    print(f"DEBUG: Chunk shape after operation {i}: {chunk.shape}")
                except Exception as e:
                    print(f"ERROR: Operation {i} failed with error: {e}")
                    print(f"ERROR: Operation details: {op}")
                    raise Exception(f"Operation {i} ({op.get('key', 'unknown')}) failed: {e}")
                    
            print(f"DEBUG: Successfully processed chunk, final shape: {chunk.shape}")
            return chunk
        except Exception as e:
            print(f"FATAL ERROR in _process_chunk: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            gc.collect()

    def save_with_operations(self, output_path: str, progress_callback=None) -> bool:
        """Apply all operations to the full file and save the result."""
        self._cancel_flag = False
        file_size = os.path.getsize(self.full_file_path)
        chunk_size = calculate_optimal_chunk_size(file_size)
        
        try:
            chunk_iterator = ChunkIterator(self.full_file_path, chunk_size)
            total_rows = chunk_iterator.total_rows
            processed_rows = 0

            if progress_callback:
                progress_callback(0, f"Starting file processing ({total_rows:,} total rows)...")

            # For CSV output
            if output_path.lower().endswith('.csv'):
                first_chunk = True
                for chunk in chunk_iterator:
                    if self._cancel_flag:
                        return False

                    # Process chunk
                    processed_chunk = self._process_chunk(chunk)
                    if processed_chunk is None:
                        return False

                    # Write chunk to CSV
                    processed_chunk.to_csv(
                        output_path,
                        mode='w' if first_chunk else 'a',
                        header=first_chunk,
                        index=False
                    )

                    processed_rows += len(chunk)
                    if progress_callback:
                        progress = processed_rows / total_rows
                        progress_callback(
                            progress,
                            f"Processed {processed_rows:,} of {total_rows:,} rows ({progress*100:.1f}%)..."
                        )

                    first_chunk = False
                    del chunk, processed_chunk
                    gc.collect()

            # For Excel output - highly optimized version with fast save
            else:
                import tempfile
                import shutil
                
                # Create a temporary directory for faster disk I/O
                with tempfile.TemporaryDirectory(dir=os.path.dirname(output_path)) as tmpdir:
                    temp_path = os.path.join(tmpdir, 'temp.xlsx')
                    
                    # Create Excel writer with optimized settings
                    workbook_options = {
                        'constant_memory': True,
                        'strings_to_numbers': False,
                        'use_zip64': True,
                        'in_memory': False,  # Write directly to disk for better memory usage
                        'default_row_height': 15,
                        'optimization': 1,
                        'tmpdir': tmpdir  # Use the same temp directory
                    }
                    
                    writer = pd.ExcelWriter(
                        temp_path,
                        engine='xlsxwriter',
                        engine_kwargs={'options': workbook_options}
                    )
                    
                    workbook = writer.book
                    worksheet = workbook.add_worksheet('Sheet1')
                    
                    # Pre-allocate format objects
                    default_format = workbook.add_format({
                        'num_format': '@',
                        'text_wrap': False
                    })
                    invalid_format = workbook.add_format({
                        'bg_color': '#FFCCCC',
                        'font_color': '#000000',
                        'num_format': '@',
                        'text_wrap': False
                    })
                    modified_format = workbook.add_format({
                        'bg_color': '#FFFFCC',
                        'font_color': '#000000',
                        'num_format': '@',
                        'text_wrap': False
                    })
                    
                    # Process chunks and write directly
                    current_row = 0
                    headers_written = False
                    
                    # Buffer for batch writing
                    write_buffer = []
                    BUFFER_SIZE = 10000  # Number of cells to buffer before writing
                    
                    def flush_buffer():
                        nonlocal write_buffer
                        if write_buffer:
                            for row, col, value, fmt in write_buffer:
                                worksheet.write(row, col, value, fmt)
                            write_buffer = []
                    
                    for chunk in chunk_iterator:
                        if self._cancel_flag:
                            return False
                            
                        # Process chunk
                        processed_chunk = self._process_chunk(chunk)
                        if processed_chunk is None:
                            return False
                        
                        # Write headers if not written yet
                        if not headers_written:
                            for col_idx, col_name in enumerate(processed_chunk.columns):
                                worksheet.write_string(0, col_idx, str(col_name), default_format)
                            current_row = 1
                            headers_written = True
                        
                        # Get numpy array of values for faster access
                        chunk_values = processed_chunk.values
                        chunk_rows, chunk_cols = chunk_values.shape
                        
                        # Pre-process styling information
                        style_masks = {}
                        modified_masks = {}
                        if hasattr(processed_chunk, '_styled_columns'):
                            for col_name, mask in processed_chunk._styled_columns.items():
                                if col_name in processed_chunk.columns:
                                    col_idx = processed_chunk.columns.get_loc(col_name)
                                    style_masks[col_idx] = mask.values
                        
                        if hasattr(processed_chunk, '_modified_columns'):
                            for col_name, mask in processed_chunk._modified_columns.items():
                                if col_name in processed_chunk.columns:
                                    col_idx = processed_chunk.columns.get_loc(col_name)
                                    modified_masks[col_idx] = mask.values
                        
                        # Write data row by row for better memory efficiency
                        for row_idx in range(chunk_rows):
                            row_data = chunk_values[row_idx]
                            for col_idx in range(chunk_cols):
                                value = row_data[col_idx]
                                is_invalid = (col_idx in style_masks and style_masks[col_idx][row_idx])
                                is_modified = (col_idx in modified_masks and modified_masks[col_idx][row_idx])
                                
                                # Choose format based on cell state (invalid takes priority)
                                if is_invalid:
                                    fmt = invalid_format
                                elif is_modified:
                                    fmt = modified_format
                                else:
                                    fmt = default_format
                                
                                write_buffer.append([current_row + row_idx, col_idx, str(value), fmt])
                                if len(write_buffer) >= BUFFER_SIZE:
                                    flush_buffer()
                        
                        # Update position and progress
                        current_row += chunk_rows
                        processed_rows += chunk_rows
                        
                        if progress_callback:
                            progress = processed_rows / total_rows
                            progress_callback(
                                progress,
                                f"Processed {processed_rows:,} of {total_rows:,} rows ({progress*100:.1f}%)..."
                            )
                        
                        # Clean up
                        del processed_chunk, chunk_values
                        if style_masks:
                            del style_masks
                        if modified_masks:
                            del modified_masks
                        gc.collect()
                    
                    # Flush any remaining data
                    flush_buffer()
                    
                    if progress_callback:
                        progress_callback(0.95, "Saving Excel file...")
                    
                    # Close the workbook to ensure all data is written
                    writer.close()
                    
                    # Move the temporary file to the final destination
                    if progress_callback:
                        progress_callback(0.98, "Moving file to final location...")
                    shutil.move(temp_path, output_path)

            if progress_callback:
                progress_callback(1.0, f"Complete! Processed {total_rows:,} rows.")

            return True

        except Exception as e:
            print(f"Error during save operation: {e}")
            raise