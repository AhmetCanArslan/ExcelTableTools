import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog, scrolledtext
import pandas as pd
import os
import re
import json
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# Add the project root to the Python path to allow imports from anywhere
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

# Add the src directory to the Python path
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

from operations.delayed_operations import DelayedOperationManager

# Import operations
from operations.masking import mask_data, mask_email, mask_words  
from operations.trimming import trim_spaces
from operations.splitting import apply_split_surname, apply_split_by_delimiter
from operations.case_change import change_case
from operations.find_replace import find_replace
from operations.remove_chars import remove_chars
from operations.concatenate import apply_concatenate
from operations.extract_pattern import apply_extract_pattern
from operations.fill_missing import fill_missing
from operations.duplicates import apply_mark_duplicates, apply_remove_duplicates
from operations.merge_columns import apply_merge_columns
from operations.rename_column import apply_rename_column   
from operations.preview_utils import generate_preview
from operations.numeric_operations import apply_round_numbers
from operations import numeric_operations
from operations.validate_inputs import apply_validation

# Import or define apply_operation_to_partition
from operations.preview_utils import apply_operation_to_partition

# Import translations
from translations import LANGUAGES

# Constants
PREVIEW_ROWS = 1000  # Number of rows to show in preview
RESOURCES_DIR = os.path.join(project_root, 'resources')

# --- GUI Application ---
class ExcelEditorApp:
    def __init__(self, root):
        self.root = root
        self.root.geometry("800x550")  
        self.file_path = tk.StringVar()
        self.selected_column = tk.StringVar()
        self.selected_operation = tk.StringVar()
        self.preview_position = tk.StringVar(value="head")
        self.dataframe = None
        self.cell_styles = None  # (row, col): {'fill':..., 'font':...}
        self.operation_manager = DelayedOperationManager()

        # --- language selection & persistence ---
        self.available_languages = list(LANGUAGES.keys())
        self.current_lang = tk.StringVar(value=self.load_last_language())
        self.texts = LANGUAGES[self.current_lang.get()]

        # --- Undo/Redo History ---
        self.undo_stack = []
        self.redo_stack = []

        # Remember last browsing directory - now uses persistent storage
        self.last_dir = self.load_last_directory()

        # load operations configuration instead of hard‐coding
        config_path = os.path.join(RESOURCES_DIR, 'operations_config.json')
        with open(config_path, "r") as f:
            self.ops_config = json.load(f)
        self.operation_keys = self.ops_config["operations"]

        # --- Main Content Frame ---
        main_content_frame = ttk.Frame(root)
        main_content_frame.pack(fill="both", expand=True, side=tk.TOP)

        # --- Top Frame for Language Selector ---
        top_frame = ttk.Frame(main_content_frame)
        top_frame.pack(fill="x", padx=10, pady=(5, 0))

        self.refresh_button = ttk.Button(top_frame, text=self.texts['refresh'], command=self.refresh_app)
        self.refresh_button.pack(side="left")

        self.lang_label = ttk.Label(top_frame, text=self.texts['language'] + ":")
        self.lang_label.pack(side="left", padx=(10,5))
        self.lang_combobox = ttk.Combobox(
            top_frame,
            textvariable=self.current_lang,
            values=self.available_languages,
            state="readonly",
            width=5
        )
        self.lang_combobox.pack(side="left")
        self.lang_combobox.bind("<<ComboboxSelected>>", lambda e: self.change_language())

        # --- File Selection ---
        self.file_frame = ttk.LabelFrame(main_content_frame, text=self.texts['file_selection'])
        self.file_frame.pack(padx=10, pady=10, fill="x")

        self.file_label = ttk.Label(self.file_frame, text=self.texts['excel_file'])
        self.file_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.file_entry = ttk.Entry(self.file_frame, textvariable=self.file_path, width=70, state="readonly")
        self.file_entry.grid(row=0, column=1, padx=5, pady=5)
        self.browse_button = ttk.Button(self.file_frame, text=self.texts['browse'], command=self.browse_file)
        self.browse_button.grid(row=0, column=2, padx=5, pady=5)

        # Add preview control frame after file selection
        preview_frame = ttk.Frame(self.file_frame)
        preview_frame.grid(row=1, column=0, columnspan=3, padx=5, pady=5)

        ttk.Label(preview_frame, text=self.texts['position']).pack(side=tk.LEFT, padx=5)
        position_combo = ttk.Combobox(preview_frame, textvariable=self.preview_position,
                                    values=[self.texts['head'], self.texts['middle'], self.texts['tail']], 
                                    state="readonly", width=8)
        position_combo.pack(side=tk.LEFT, padx=5)

        refresh_preview_btn = ttk.Button(preview_frame, text=self.texts['refresh_preview'],
                                       command=self.refresh_preview)
        refresh_preview_btn.pack(side=tk.LEFT, padx=5)

        # Bind preview control changes
        position_combo.bind('<<ComboboxSelected>>', lambda e: self.refresh_preview())

        # --- Operations ---
        self.ops_frame = ttk.LabelFrame(main_content_frame, text=self.texts['operations'])
        self.ops_frame.pack(padx=10, pady=10, fill="x")

        self.column_label = ttk.Label(self.ops_frame, text=self.texts['column'])
        self.column_label.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.column_combobox = ttk.Combobox(self.ops_frame, textvariable=self.selected_column, state="disabled", width=85)
        self.column_combobox.grid(row=0, column=1, padx=5, pady=5)

        self.operation_label = ttk.Label(self.ops_frame, text=self.texts['operation'])
        self.operation_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.operation_combobox = ttk.Combobox(self.ops_frame, textvariable=self.selected_operation, state="disabled", width=85)
        self.operation_combobox.grid(row=1, column=1, padx=5, pady=5)

        # Create a frame to hold buttons
        buttons_frame = ttk.Frame(self.ops_frame)
        buttons_frame.grid(row=2, column=0, columnspan=2, padx=5, pady=10, sticky="ew")
        
        # Split the frame into a 3-column grid
        buttons_frame.columnconfigure(0, weight=1)
        buttons_frame.columnconfigure(1, weight=1)

        self.apply_button = ttk.Button(buttons_frame, text=self.texts['apply_operation'], command=self.apply_operation)
        self.apply_button.grid(row=0, column=0, padx=2, sticky="ew")

        self.preview_button = ttk.Button(buttons_frame,
            text=self.texts.get('operation_preview_button', "Operation Preview"),
            command=self.preview_operation,
            state="disabled")  # start disabled
        self.preview_button.grid(row=0, column=1, padx=2, sticky="ew")
        
        self.output_preview_button = ttk.Button(buttons_frame,
            text=self.texts.get('output_preview_button', "Output File Preview"),
            command=self.preview_output_file,
            state="disabled")
        self.output_preview_button.grid(row=0, column=2, padx=2, sticky="ew")

        # toggle preview buttons when operation selection changes
        self.selected_operation.trace_add("write", self._on_operation_change)
        
        # When a file is loaded, enable output preview
        self.file_path.trace_add("write", self._on_file_loaded)

        self.ops_frame.columnconfigure(0, weight=1)
        self.ops_frame.columnconfigure(1, weight=1)

        # --- Save and Undo/Redo Frame ---
        save_frame = ttk.Frame(main_content_frame)
        save_frame.pack(padx=10, pady=10, fill="x")

        self.undo_button = ttk.Button(save_frame, text="Undo", command=self.undo_action, state="disabled")
        self.undo_button.pack(side="left", padx=5)

        self.redo_button = ttk.Button(save_frame, text="Redo", command=self.redo_action, state="disabled")
        self.redo_button.pack(side="left", padx=5)

        # Add a dropdown to choose output file extension
        self.output_extension = tk.StringVar(value="xlsx")  # Set default value
        self.output_formats = ["xlsx", "xls", "csv", "json", "html", "md"]
        
        ttk.Label(save_frame, text=self.texts['output_format']).pack(side="right", padx=(5,0))
        self.extension_dropdown = ttk.Combobox(save_frame, textvariable=self.output_extension,
                                             values=self.output_formats, state="readonly", width=5)
        self.extension_dropdown.pack(side="right", padx=5)
        # Bind the dropdown selection event
        self.extension_dropdown.bind('<<ComboboxSelected>>', self._on_extension_change)

        self.save_button = ttk.Button(save_frame, text=self.texts['save_changes'], command=self.save_file)
        self.save_button.pack(side="right", padx=5)

        # --- Status Area (CLI-like) ---
        self.status_frame = ttk.LabelFrame(root, text=self.texts['status_log'])
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(5, 10))

        self.status_text = scrolledtext.ScrolledText(self.status_frame, height=5, wrap=tk.WORD, state='disabled')
        self.status_text.pack(fill="both", expand=True, padx=5, pady=5)

        self.update_status(self.texts['ready'])

        self.update_ui_language()

    def get_unique_col_name(self, base_name, existing_columns):
        """Generates a unique column name based on existing ones."""
        new_name = base_name
        counter = 1
        while new_name in existing_columns:
            new_name = f"{base_name}_{counter}"
            counter += 1
        return new_name

    def update_status(self, message):
        """Appends a message to the status text area."""
        self.status_text.config(state='normal')
        self.status_text.insert(tk.END, message + "\n")
        self.status_text.see(tk.END)
        self.status_text.config(state='disabled')

    def update_ui_language(self):
        """Updates all UI elements with the current language."""
        self.texts = LANGUAGES[self.current_lang.get()]
        self.root.title(self.texts['title'])

        # Update frame labels
        self.file_frame.config(text=self.texts['file_selection'])
        self.ops_frame.config(text=self.texts['operations'])
        self.status_frame.config(text=self.texts['status_log'])

        # Update labels and buttons
        self.file_label.config(text=self.texts['excel_file'])
        self.browse_button.config(text=self.texts['browse'])
        self.column_label.config(text=self.texts['column'])
        self.operation_label.config(text=self.texts['operation'])
        self.apply_button.config(text=self.texts['apply_operation'])
        self.preview_button.config(text=self.texts['operation_preview_button'])
        self.output_preview_button.config(text=self.texts['output_preview_button'])
        self.save_button.config(text=self.texts['save_changes'])
        self.refresh_button.config(text=self.texts['refresh'])
        self.lang_label.config(text=self.texts['language'] + ":")
        
        # Update preview controls
        for widget in self.file_frame.winfo_children():
            if isinstance(widget, ttk.Frame):  # This is our preview_frame
                for child in widget.winfo_children():
                    if isinstance(child, ttk.Label):
                        if child.cget("text").startswith("Position"):
                            child.config(text=self.texts['position'])
                    elif isinstance(child, ttk.Button):
                        if child.cget("text").startswith("Refresh"):
                            child.config(text=self.texts['refresh_preview'])
                    elif isinstance(child, ttk.Combobox):
                        current_val = child.get()
                        new_values = [self.texts['head'], self.texts['middle'], self.texts['tail']]
                        child.config(values=new_values)
                        # Map old value to new translated value
                        value_map = {'head': 'head', 'middle': 'middle', 'tail': 'tail'}
                        if current_val.lower() in value_map:
                            child.set(self.texts[value_map[current_val.lower()]])

        # Update Undo/Redo buttons
        self.undo_button.config(text=self.texts['undo'])
        self.redo_button.config(text=self.texts['redo'])

        # Update operations combobox
        translated_ops = [self.texts[key] for key in self.operation_keys]
        current_selection_text = self.selected_operation.get()
        self.operation_combobox['values'] = translated_ops

        if current_selection_text:
            try:
                current_key = None
                old_lang = 'tr' if self.current_lang.get() == 'en' else 'en'
                for key, text in LANGUAGES[old_lang].items():
                    if text == current_selection_text and key in self.operation_keys:
                        current_key = key
                        break
                if current_key:
                    self.selected_operation.set(self.texts[current_key])
                else:
                    self.selected_operation.set("")
            except Exception:
                self.selected_operation.set("")
        else:
            self.selected_operation.set("")

        # Update initial status
        if not self.dataframe is None:
            self.update_status(self.texts['ready'])

    def load_last_language(self):
        """Load last selected language or default to 'en'."""
        try:
            lang_file_path = os.path.join(RESOURCES_DIR, "last_language.txt")
            if os.path.exists(lang_file_path):
                with open(lang_file_path, "r") as f:
                    lang = f.read().strip()
                    return lang if lang in LANGUAGES else 'en'
            return 'en'
        except Exception:
            return 'en'

    def save_last_language(self):
        """Save current language to file."""
        try:
            lang_file_path = os.path.join(RESOURCES_DIR, "last_language.txt")
            with open(lang_file_path, "w") as f:
                f.write(self.current_lang.get())
        except Exception:
            pass
            
    def load_last_directory(self):
        """Load last browsing directory or default to current directory."""
        try:
            dir_file_path = os.path.join(RESOURCES_DIR, "last_directory.txt")
            if os.path.exists(dir_file_path):
                with open(dir_file_path, "r") as f:
                    directory = f.read().strip()
                    return directory if os.path.isdir(directory) else os.getcwd()
            return os.getcwd()
        except Exception:
            return os.getcwd()
            
    def save_last_directory(self):
        """Save current browsing directory to file."""
        try:
            dir_file_path = os.path.join(RESOURCES_DIR, "last_directory.txt")
            with open(dir_file_path, "w") as f:
                f.write(self.last_dir)
        except Exception:
            pass

    def change_language(self):
        """Apply and persist the selected language."""
        lang = self.current_lang.get()
        self.save_last_language()
        self.texts = LANGUAGES[lang]
        self.update_ui_language()

    def refresh_app(self):
        """Resets the application to its initial state."""
        # Clear file and data
        self.file_path.set("")
        self.dataframe = None
        self.original_df = None  # Ensure original_df is cleared
        
        # Clear operation manager state
        self.operation_manager.clear_operations()  # Clear pending operations
        if hasattr(self.operation_manager, 'full_file_path'):
            self.operation_manager.full_file_path = None
        
        # Disable & clear comboboxes
        self.column_combobox.set("")
        self.column_combobox['values'] = []
        self.column_combobox.config(state="disabled")
        self.operation_combobox.set("")
        self.operation_combobox['values'] = []
        self.operation_combobox.config(state="disabled")
        self.operation_combobox['values'] = [self.texts[key] for key in self.operation_keys]

        # Clear undo/redo history
        self.undo_stack.clear()
        self.redo_stack.clear()
        self.update_undo_redo_buttons()

        # Clear any styling information
        if hasattr(self, 'cell_styles'):
            self.cell_styles = None

        # Clear status log
        self.status_text.config(state='normal')
        self.status_text.delete('1.0', tk.END)
        self.status_text.config(state='disabled')

        # Reset output extension dropdown
        self.output_extension.set("xlsx")

        # disable preview buttons
        self.preview_button.config(state="disabled")
        self.output_preview_button.config(state="disabled")

        # Force garbage collection to clean up memory
        import gc
        gc.collect()

        # Inform user
        self.update_status(self.texts['app_refreshed'])

    def get_operation_key(self, translated_op_text):
        for key in self.operation_keys:
            if self.texts[key] == translated_op_text:
                return key
        return None

    def _on_operation_change(self, *args):
        """Enable Preview only if a valid operation is selected."""
        op_text = self.selected_operation.get()
        if self.get_operation_key(op_text):
            self.preview_button.config(state="normal")
        else:
            self.preview_button.config(state="disabled")
            
    def _on_file_loaded(self, *args):
        """Enable output preview button when a file is loaded."""
        if self.file_path.get():
            self.output_preview_button.config(state="normal")
        else:
            self.output_preview_button.config(state="disabled")

    def browse_file(self):
        path = filedialog.askopenfilename(
            initialdir=self.last_dir,
            title=self.texts['select_excel_file'],
            filetypes=[(self.texts['excel_files'], "*.xlsx *.xls *.csv")]
        )
        if path:
            self.last_dir = os.path.dirname(path)
            self.save_last_directory()
            self.file_path.set(path)
            self.load_preview()
        else:
            self.update_status(self.texts['file_selection_cancelled'])

    def load_preview(self):
        path = self.file_path.get()
        if not path:
            return
        try:
            # Load preview using operation manager
            self.dataframe = self.operation_manager.load_preview(
                path,
                self.preview_position.get()
            )

            # Store original dataframe when first loading
            if not hasattr(self, 'original_df'):
                self.original_df = self.dataframe.copy(deep=True)

            # Update UI
            self.column_combobox['values'] = list(self.dataframe.columns)
            self.column_combobox.config(state="readonly")
            if self.dataframe.columns.any():
                self.selected_column.set(self.dataframe.columns[0])
            self.operation_combobox.config(state="readonly")
            
            preview_type = f"1000 rows from {self.preview_position.get()}"
            messagebox.showinfo(
                self.texts['success'],
                f"Loaded preview ({preview_type}) from '{os.path.basename(path)}'"
            )
            self.update_status(self.texts['loaded_preview'].format(type=preview_type))

        except Exception as e:
            messagebox.showerror(self.texts['error'], self.texts['error_loading'].format(error=e))
            self.file_path.set("")
            self.dataframe = None
            self.original_df = None  # Also clear original_df on error
            self.column_combobox['values'] = []
            self.column_combobox.config(state="disabled")
            self.operation_combobox.config(state="disabled")
            self.selected_column.set("")
            self.selected_operation.set("")
            self.update_status(self.texts['error_loading_file'].format(error=str(e)))

    def refresh_preview(self):
        if self.file_path.get():
            self.load_preview()

    def preview_operation(self):
        """Show side-by-side preview: original loaded sample vs current output sample."""
        if self.dataframe is None or self.dataframe.empty:
            messagebox.showwarning(
                self.texts['warning'],
                self.texts['preview_no_data'],
                parent=self.root
            )
            self.update_status(
                self.texts['preview_status_message'].format(message=self.texts['preview_no_data'])
            )
            return

        # Get original DataFrame, or use current DataFrame if original is not available
        orig = getattr(self, 'original_df', None)
        if orig is None:
            orig = self.dataframe
            self.original_df = orig.copy(deep=True)  # Store it for future use
            
        original_sample = orig.head(PREVIEW_ROWS).copy(deep=True)

        # Get operation details
        op_text = self.selected_operation.get()
        op_key = self.get_operation_key(op_text)
        col = self.selected_column.get()
        
        # Handle operations that require user input
        operation_params = {
            'type': 'column_operation',
            'key': op_key,
            'column': col
        }
        
        # Get user input based on operation type
        if op_key == 'op_find_replace':
            find_text = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_find_text'],
                parent=self.root
            )
            if find_text is None:  # User cancelled
                return
                
            replace_text = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_replace_text'],
                parent=self.root
            )
            if replace_text is None:  # User cancelled
                return
                
            operation_params['find_text'] = find_text
            operation_params['replace_text'] = replace_text
            
        elif op_key == 'op_split_delimiter':
            delimiter = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_delimiter'],
                parent=self.root
            )
            if delimiter is None:  # User cancelled
                return
                
            operation_params['delimiter'] = delimiter
            
        elif op_key == 'op_remove_specific':
            chars = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_chars_to_remove'],
                parent=self.root
            )
            if chars is None:  # User cancelled
                return
                
            operation_params['chars_to_remove'] = chars
            
        elif op_key == 'op_fill_missing':
            fill_value = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_fill_value'],
                parent=self.root
            )
            if fill_value is None:  # User cancelled
                return
                
            operation_params['fill_value'] = fill_value
            
        elif op_key == 'op_extract_pattern':
            pattern = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_regex_pattern'],
                parent=self.root
            )
            if pattern is None:  # User cancelled
                return
                
            new_col_name = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_new_col_name'],
                parent=self.root
            )
            if new_col_name is None:  # User cancelled
                return
                
            operation_params['pattern'] = pattern
            operation_params['new_col_name'] = new_col_name

        elif op_key == 'op_split_surname':
            # No additional input needed for surname splitting
            pass

        # Use generate_preview to apply the operation preview
        modified_sample = self.dataframe.head(PREVIEW_ROWS).copy(deep=True)
        
        if op_key:
            preview_df, success, msg = generate_preview(self, op_key, self.selected_column.get(), modified_sample, PREVIEW_ROWS, operation_params)
            if success and preview_df is not None:
                modified_sample = preview_df
            else:
                self.update_status(self.texts.get('preview_failed', "Preview failed: {error}").format(error=msg))
                return

        # Preserve styling information (if present) using a safer approach
        if hasattr(self.dataframe, '_styled_columns'):
            object.__setattr__(modified_sample, '_styled_columns', {})
            for col, mask in self.dataframe._styled_columns.items():
                if col in modified_sample.columns:
                    modified_sample._styled_columns[col] = mask.head(PREVIEW_ROWS).copy()

        self.show_preview_dialog(
            original_sample,
            modified_sample,
            self.texts.get('preview_output_title', "Output Preview")
        )

    def preview_output_file(self):
        """Show preview of the current state of the dataframe (all operations applied)."""
        if self.dataframe is None or self.dataframe.empty:
            messagebox.showwarning(
                self.texts['warning'],
                self.texts.get('preview_no_data', "No data to preview."),
                parent=self.root
            )
            self.update_status("Output file preview failed: No data available.")
            return

        # Get a sample of the original file for comparison
        if hasattr(self, 'original_df') and self.original_df is not None:
            original_sample = self.original_df.head(PREVIEW_ROWS).copy(deep=True)
        else:
            # If no original_df, use current dataframe as both original and modified
            original_sample = self.dataframe.head(PREVIEW_ROWS).copy(deep=True)
            self.original_df = self.dataframe.copy(deep=True)
            
        # Get a sample of the current state with all operations applied
        current_sample = self.dataframe.head(PREVIEW_ROWS).copy(deep=True)
            
        # Preserve styling information if present
        if hasattr(self.dataframe, '_styled_columns'):
            object.__setattr__(current_sample, '_styled_columns', {})
            for col, mask in self.dataframe._styled_columns.items():
                if col in current_sample.columns:
                    current_sample._styled_columns[col] = mask.head(PREVIEW_ROWS).copy()
                
        # Show the preview dialog comparing original to current state
        self.show_preview_dialog(
            original_sample,
            current_sample,
            self.texts.get('output_file_preview_title', "Current Output File")
        )
        
        # Display summary of changes in status log
        original_cols = set(original_sample.columns)
        current_cols = set(current_sample.columns)
        added_cols = current_cols - original_cols
        removed_cols = original_cols - current_cols
        
        summary_parts = []
        
        # Report on column changes
        if added_cols:
            summary_parts.append(f"Added columns: {', '.join(sorted(added_cols))}")
        if removed_cols:
            summary_parts.append(f"Removed columns: {', '.join(sorted(removed_cols))}")
            
        # Report on row changes
        orig_rows = len(self.original_df) if hasattr(self, 'original_df') and self.original_df is not None else "unknown"
        current_rows = len(self.dataframe)
        if orig_rows != current_rows:
            summary_parts.append(f"Rows changed: {orig_rows} → {current_rows}")
        
        # Report on output format
        output_format = self.output_extension.get()
        summary_parts.append(f"Output format: {output_format.upper()}")
        
        # Log the summary
        if summary_parts:
            summary = " | ".join(summary_parts)
            self.update_status(self.texts['output_preview_summary'].format(summary=summary))
            
            
    # --- Undo/Redo Methods ---
    def _commit_undoable_action(self, old_df):
        """Saves the current 'old' DataFrame in undo stack and clears redo stack."""
        self.undo_stack.append(old_df)
        self.redo_stack.clear()
        self.update_undo_redo_buttons()

    def update_undo_redo_buttons(self):
        """Enable or disable Undo/Redo buttons based on stack states."""
        if self.undo_stack:
            self.undo_button.config(state="normal")
        else:
            self.undo_button.config(state="disabled")

        if self.redo_stack:
            self.redo_button.config(state="normal")
        else:
            self.redo_button.config(state="disabled")

    def undo_action(self):
        if not self.undo_stack:
            messagebox.showwarning(self.texts['warning'], self.texts['nothing_to_undo'], parent=self.root)
            return

        # Store current state for redo
        current_df = self.dataframe.copy()
        self.redo_stack.append(current_df)

        # Restore from undo stack
        self.dataframe = self.undo_stack.pop()

        # Remove last operation from queue
        if self.operation_manager.operations:
            self.operation_manager.operations.pop()

        # Update UI
        self.update_column_combobox()
        self.update_undo_redo_buttons()
        self.update_status(self.texts['undo_performed'])

    def redo_action(self):
        if not self.redo_stack:
            messagebox.showwarning(self.texts['warning'], self.texts['nothing_to_redo'], parent=self.root)
            return

        # Store current state for undo
        current_df = self.dataframe.copy()
        self.undo_stack.append(current_df)

        # Restore from redo stack
        self.dataframe = self.redo_stack.pop()

        # Re-add the operation to queue if available
        if hasattr(self, '_last_undone_operation'):
            self.operation_manager.operations.append(self._last_undone_operation)

        # Update UI
        self.update_column_combobox()
        self.update_undo_redo_buttons()
        self.update_status(self.texts['redo_performed'])

    def apply_operation(self):
        if self.dataframe is None:
            self.update_status(self.texts['operation_failed_no_file'])
            return

        col = self.selected_column.get()
        op_text = self.selected_operation.get()
        op_key = self.get_operation_key(op_text)

        # Store original state for undo
        original_df = self.dataframe.copy()

        # Ensure we have the original DataFrame stored
        if not hasattr(self, 'original_df'):
            self.original_df = original_df.copy(deep=True)

        # Create operation object with only serializable data
        operation = {
            'type': 'column_operation',
            'key': op_key,
            'column': col
        }

        # Handle operations that require user input
        if op_key == 'op_find_replace':
            find_text = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_find_text'],
                parent=self.root
            )
            if find_text is None:  # User cancelled
                return
                
            replace_text = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_replace_text'],
                parent=self.root
            )
            if replace_text is None:  # User cancelled
                return
                
            operation['find_text'] = find_text
            operation['replace_text'] = replace_text
            
        elif op_key == 'op_split_delimiter':
            delimiter = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_delimiter'],
                parent=self.root
            )
            if delimiter is None:  # User cancelled
                return
                
            operation['delimiter'] = delimiter
            
        elif op_key == 'op_remove_specific':
            chars = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_chars_to_remove'],
                parent=self.root
            )
            if chars is None:  # User cancelled
                return
                
            operation['chars_to_remove'] = chars
            
        elif op_key == 'op_fill_missing':
            fill_value = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_fill_value'],
                parent=self.root
            )
            if fill_value is None:  # User cancelled
                return
                
            operation['fill_value'] = fill_value
            
        elif op_key == 'op_extract_pattern':
            pattern = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_regex_pattern'],
                parent=self.root
            )
            if pattern is None:  # User cancelled
                return
                
            new_col_name = simpledialog.askstring(
                self.texts['input_needed'],
                self.texts['enter_new_col_name'],
                parent=self.root
            )
            if new_col_name is None:  # User cancelled
                return
                
            operation['pattern'] = pattern
            operation['new_col_name'] = new_col_name

        elif op_key == 'op_split_surname':
            # No additional input needed for surname splitting
            pass

        # Add operation to manager for full processing later
        self.operation_manager.add_operation(operation)

        # Apply operation immediately to preview data
        try:
            # Apply operation directly to preview data
            self.dataframe = apply_operation_to_partition(self.dataframe, operation['type'], operation)
            
            # Add to undo stack and clear redo stack
            self.undo_stack.append(original_df)
            self.redo_stack.clear()
            self.update_undo_redo_buttons()
            
            # Update UI
            self.update_column_combobox()
            
            # Update status log
            self.update_status(self.texts['added_operation'].format(
                operation=op_text,
                column=col
            ))
        except Exception as e:
            # Restore original state on error
            self.dataframe = original_df
            messagebox.showerror(
                self.texts['error'],
                f"Error applying operation to preview: {str(e)}"
            )
            self.update_status(self.texts['preview_error'].format(error=str(e)))
            # Remove the operation from queue if preview failed
            self.operation_manager.operations.pop()

    def show_preview_dialog(self, original_df_sample, modified_df_sample, op_text):
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title(self.texts['preview_display_title'])
        preview_dialog.transient(self.root)
        preview_dialog.grab_set()

        width = 1000
        height = 700
        preview_dialog.geometry(f"{width}x{height}")
        preview_dialog.resizable(True, True)    

        main_frame = ttk.Frame(preview_dialog, padding="10")
        main_frame.pack(expand=True, fill="both")

        ttk.Label(main_frame, text=f"{op_text} - {self.texts['preview_display_title']}").pack(pady=5)

        notebook = ttk.Notebook(main_frame)
        notebook.pack(expand=True, fill="both", pady=5)

        # Format data as a well-aligned table string
        def format_dataframe_as_table(df):
            table_string = ""
            
            # Get maximum width for each column for proper alignment
            col_widths = {'row_num': 5}  # Start with row numbers column
            for col in df.columns:
                # Get max width of column name and values
                col_values = df[col].astype(str)
                max_value_width = max((col_values.str.len().max(), len(str(col))))
                col_widths[col] = max_value_width + 3  # Add padding
            
            # Create header
            header_row = f"{'#':<5}"  # Row number header
            separator_row = "-" * 5 + " "
            for col in df.columns:
                width = col_widths[col]
                header_row += f"{str(col):<{width}}"
                separator_row += "-" * width + " "
            
            table_string += header_row + "\n"
            table_string += separator_row + "\n"
            
            # Create data rows
            for idx, (_, row) in enumerate(df.iterrows(), 1):
                data_row = f"{idx:<5}"  # Add row number
                for col in df.columns:
                    width = col_widths[col]
                    value = str(row[col]) if not pd.isna(row[col]) else ""
                    data_row += f"{value:<{width}}"
                table_string += data_row + "\n"
                
            return table_string

        # Create the original data tab
        original_tab = ttk.Frame(notebook)
        notebook.add(original_tab, text=self.texts['preview_original_data'].format(n=PREVIEW_ROWS))
        
        # Create frame for original data with scrollbars
        original_frame = ttk.Frame(original_tab)
        original_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Add Text widget with both scrollbars
        original_text = tk.Text(original_frame, wrap=tk.NONE, font=("Courier", 15))
        original_text.insert(tk.END, format_dataframe_as_table(original_df_sample))
        original_text.config(state="disabled")
        
        # Create and position scrollbars
        original_vsb = ttk.Scrollbar(original_frame, orient="vertical", command=original_text.yview)
        original_hsb = ttk.Scrollbar(original_frame, orient="horizontal", command=original_text.xview)
        original_text.configure(yscrollcommand=original_vsb.set, xscrollcommand=original_hsb.set)
        
        # Grid layout for text and scrollbars
        original_frame.grid_rowconfigure(0, weight=1)
        original_frame.grid_columnconfigure(0, weight=1)
        original_text.grid(row=0, column=0, sticky="nsew")
        original_vsb.grid(row=0, column=1, sticky="ns")
        original_hsb.grid(row=1, column=0, sticky="ew")
        
        # Create the modified data tab
        modified_tab = ttk.Frame(notebook)
        notebook.add(modified_tab, text=self.texts['preview_modified_data'].format(n=PREVIEW_ROWS))
        
        # Create frame for modified data with scrollbars
        modified_frame = ttk.Frame(modified_tab)
        modified_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Add Text widget with both scrollbars
        modified_text = tk.Text(modified_frame, wrap=tk.NONE, font=("Courier", 15))
        
        # If there is styling information, apply it using tags
        has_styling = hasattr(modified_df_sample, '_styled_columns')
        has_modifications = hasattr(modified_df_sample, '_modified_columns')
        
        if has_styling or has_modifications:
            # Configure tags for different cell states
            modified_text.tag_configure("invalid", background="#FFCCCC")
            modified_text.tag_configure("modified", background="#FFFFCC")
            
            # Format table with styling
            styled_table = format_dataframe_as_table(modified_df_sample)
            modified_text.insert(tk.END, styled_table)
            
            # Apply tags for styled cells (invalid)
            if has_styling:
                for col, mask in modified_df_sample._styled_columns.items():
                    if col in modified_df_sample.columns:
                        col_idx = list(modified_df_sample.columns).index(col)
                        
                        for row_idx, is_invalid in enumerate(mask):
                            if is_invalid:
                                line_num = row_idx + 2  # +2 for header and separator
                                pos = 5  # Start after row number column
                                for j, c in enumerate(modified_df_sample.columns):
                                    if j == col_idx:
                                        break
                                    col_width = max(len(str(c)), modified_df_sample[c].astype(str).str.len().max()) + 3
                                    pos += col_width
                                    
                                cell_value = str(modified_df_sample.iloc[row_idx, col_idx])
                                start_pos = f"{line_num + 1}.{pos}"
                                end_pos = f"{start_pos}+{len(cell_value)}c"
                                modified_text.tag_add("invalid", start_pos, end_pos)
            
            # Apply tags for modified cells (if not already invalid)
            if has_modifications:
                for col, mask in modified_df_sample._modified_columns.items():
                    if col in modified_df_sample.columns:
                        col_idx = list(modified_df_sample.columns).index(col)
                        
                        for row_idx, is_modified in enumerate(mask):
                            if is_modified:
                                # Check if this cell is already marked as invalid
                                is_already_invalid = False
                                if has_styling and col in modified_df_sample._styled_columns:
                                    is_already_invalid = modified_df_sample._styled_columns[col].iloc[row_idx]
                                
                                if not is_already_invalid:  # Only highlight as modified if not invalid
                                    line_num = row_idx + 2  # +2 for header and separator
                                    pos = 5  # Start after row number column
                                    for j, c in enumerate(modified_df_sample.columns):
                                        if j == col_idx:
                                            break
                                        col_width = max(len(str(c)), modified_df_sample[c].astype(str).str.len().max()) + 3
                                        pos += col_width
                                        
                                    cell_value = str(modified_df_sample.iloc[row_idx, col_idx])
                                    start_pos = f"{line_num + 1}.{pos}"
                                    end_pos = f"{start_pos}+{len(cell_value)}c"
                                    modified_text.tag_add("modified", start_pos, end_pos)
        else:
            # Just add the table without styling
            modified_text.insert(tk.END, format_dataframe_as_table(modified_df_sample))
            
        modified_text.config(state="disabled")
        
        # Create and position scrollbars
        modified_vsb = ttk.Scrollbar(modified_frame, orient="vertical", command=modified_text.yview)
        modified_hsb = ttk.Scrollbar(modified_frame, orient="horizontal", command=modified_text.xview)
        modified_text.configure(yscrollcommand=modified_vsb.set, xscrollcommand=modified_hsb.set)
        
        # Grid layout for text and scrollbars
        modified_frame.grid_rowconfigure(0, weight=1)
        modified_frame.grid_columnconfigure(0, weight=1)
        modified_text.grid(row=0, column=0, sticky="nsew")
        modified_vsb.grid(row=0, column=1, sticky="ns")
        modified_hsb.grid(row=1, column=0, sticky="ew")
        
        # Add note about styling if needed
        if has_styling or has_modifications:
            note_text = ""
            if has_styling:
                note_text += self.texts.get('validation_highlight_note', "Cells highlighted in red failed validation.")
            if has_modifications:
                if note_text:
                    note_text += " "
                note_text += self.texts.get('modification_highlight_note', "Cells highlighted in yellow were modified by operations.")
            
            validation_note = ttk.Label(
                modified_tab, 
                text=note_text,
                font=("", 9, "italic")
            )
            validation_note.pack(pady=(5, 0))
        elif has_styling:  # Keep the original note for backward compatibility
            validation_note = ttk.Label(
                modified_tab, 
                text=self.texts.get('validation_highlight_note', "Cells highlighted in red failed validation."),
                font=("", 9, "italic")
            )
            validation_note.pack(pady=(5, 0))
        
        # Add OK button at the bottom
        ttk.Button(main_frame, text="OK", command=preview_dialog.destroy).pack(pady=10)

    def update_column_combobox(self, preferred_selection=None):
        if self.dataframe is not None:
            cols = list(self.dataframe.columns)
            self.column_combobox['values'] = cols
            if preferred_selection and preferred_selection in cols:
                self.selected_column.set(preferred_selection)
            elif cols:
                self.selected_column.set(cols[0])
            else:
                self.selected_column.set("")
            # Eğer yeni bir sütun eklendiyse, genişliğini ayarla
            if preferred_selection and preferred_selection in cols:
                self.set_column_width_to_content(preferred_selection)
        else:
            self.column_combobox['values'] = []
            self.selected_column.set("")

    def set_column_width_to_content(self, col_name):
        """Set the column width for a given column based on its content length."""
        if self.dataframe is None or col_name not in self.dataframe.columns:
            return
        max_len = max(
            [len(str(val)) if not pd.isna(val) else 0 for val in self.dataframe[col_name]] + [len(str(col_name))]
        )
        # Excel'de yaklaşık karakter genişliği için biraz ek boşluk bırak
        width = max_len + 2
        # self.cell_styles yoksa oluştur
        if self.cell_styles is None:
            self.cell_styles = {'cell_styles': {}, 'col_widths': {}, 'row_heights': {}}
        if 'col_widths' not in self.cell_styles:
            self.cell_styles['col_widths'] = {}
        # Excel sütun harfi bul
        col_idx = list(self.dataframe.columns).index(col_name)
        col_letter = get_column_letter(col_idx + 1)
        self.cell_styles['col_widths'][col_letter] = width

    def _on_extension_change(self, event):
        """Handle extension dropdown value changes."""
        selected_extension = self.output_extension.get()
        self.update_status(self.texts['output_format_changed'].format(format=selected_extension))

    def save_file(self):
        if self.dataframe is None:
            self.update_status(self.texts['save_failed_no_data'])
            return

        # Get the selected extension from the dropdown
        output_ext = self.output_extension.get()
        
        original_name = os.path.splitext(os.path.basename(self.file_path.get()))[0]
        suggested_name = f"{original_name}_modified.{output_ext}"

        save_path = filedialog.asksaveasfilename(
            initialdir=self.last_dir,
            title=self.texts['save_modified_file'],
            initialfile=suggested_name,
            defaultextension=f".{output_ext}",
            filetypes=[
                (self.texts['excel_files'], "*.xlsx;*.xls;*.csv"),
                ("Excel Files (*.xlsx)", "*.xlsx"),
                ("Excel 97-2003 (*.xls)", "*.xls"),
                ("CSV Files (*.csv)", "*.csv"),
                ("JSON Files (*.json)", "*.json"),
                ("HTML Files (*.html)", "*.html"),
                ("Markdown Files (*.md)", "*.md")
            ]
        )

        if not save_path:
            self.update_status(self.texts['save_cancelled'])
            return

        try:
            # Create progress dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Saving File")
            progress_window.transient(self.root)
            progress_window.grab_set()

            progress_var = tk.DoubleVar()
            progress_label = ttk.Label(progress_window, text="Starting...")
            progress_label.pack(pady=5)
            
            progress_bar = ttk.Progressbar(
                progress_window,
                variable=progress_var,
                maximum=100,
                mode='determinate'
            )
            progress_bar.pack(padx=10, pady=5, fill=tk.X)

            cancel_button = ttk.Button(
                progress_window,
                text="Cancel",
                command=self.operation_manager.cancel_processing
            )
            cancel_button.pack(pady=5)

            def update_progress(progress, message):
                progress_var.set(progress * 100)
                progress_label.config(text=message)
                progress_window.update()

            # Start processing in a separate thread
            def process_file():
                try:
                    success = self.operation_manager.save_with_operations(
                        save_path,
                        progress_callback=update_progress
                    )
                    
                    progress_window.after(0, progress_window.destroy)
                    
                    if success:
                        messagebox.showinfo(
                            self.texts['success'],
                            self.texts['file_saved'].format(filename=os.path.basename(save_path))
                        )
                        self.update_status(f"File saved successfully to {os.path.basename(save_path)}.")
                    else:
                        messagebox.showwarning(
                            self.texts['warning'],
                            "Operation was cancelled or encountered an error."
                        )
                        self.update_status(self.texts['save_error_or_cancelled'])
                except Exception as e:
                    progress_window.after(0, progress_window.destroy)
                    messagebox.showerror(
                        self.texts['error'],
                        f"Error saving file: {str(e)}"
                    )
                    self.update_status(f"Error saving file: {str(e)}")

            import threading
            thread = threading.Thread(target=process_file)
            thread.start()

        except Exception as e:
            messagebox.showerror(
                self.texts['error'],
                f"Error setting up save operation: {str(e)}"
            )
            self.update_status(f"Error setting up save operation: {str(e)}")

# Add the main block to start the application
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelEditorApp(root)
    root.mainloop()